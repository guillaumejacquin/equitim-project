import site
import pymongo
from pymongo import MongoClient

from calculs.sponsor import sponsor

#ajoute une valeud dans la collection clients qui se trouve dans la base de donnees templates
def add_value_data_base(ticker, equity, dividende="", sponsor="", siteweb="", inconvénient=""):
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")

    db = cluster["templates"]
    collection = db["clients"]

    post1 = {"Ticker": ticker, "Equity" : equity, "Dividende": dividende, "Sponsor": sponsor, "SiteWeb": siteweb, "Inconvénient": inconvénient}

    collection.insert_one(post1)

# add_value_data_base("Cac50", "ca")


#montre les valeurs de la base de la collections clients de la base de données templates
def show_database():
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")

    db = cluster["templates"]
    collection = db["clients"]
    list_elements_database = []
    for x in collection.find({}, {"_id":0, "Libelle": 1, "Ticker": 1 }): 
        list_elements_database.append(x)

    return(list_elements_database)


# show_database()

from openpyxl import load_workbook

def add_value():
    file_path = 'Classeur1.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Feuil1']  # or wb.active
    
    i = 2
    while True:
        mystring = "A" + str(i)
        secondstring = "B" + str(i)

        if (ws[mystring].value is None):
            exit (2)

        if (ws[mystring].value is None):   
            ws[mystring].value = " "
        add_value_data_base(ws[mystring].value, ws[secondstring].value)
        i+=1
    wb.save(file_path)
# add_value()

def takeinformations(Class):
    #Recupere le sous jacent
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")

    db = cluster["templates"]
    collection = db["clients"]

    stringlongue = ""
    #remplacer par le bon element(ici ticker)



    myresults = []
    for i in Class.TSJ:
        myresults.append(collection.find({"Ticker":i}))

    compteur = 0
    for result in myresults:
            if compteur == 0:
                mot = ""
            else:
                mot = " et "
            try:
                test = result[0]
                Class.NOMSOUSJACENT = Class.NOMSOUSJACENT + mot + (test["Equity"])
                Class.DIVIDENDE = Class.DIVIDENDE + mot + test["Dividende"]
                Class.SPONSOR = Class.SPONSOR + mot + test["Sponsor"]
                Class.Site = Class.Site + mot + test["SiteWeb"]
                Class.TICKER = Class.TICKER + mot + test["Ticker"]
            except Exception:
                Class.NOMSOUSJACENT + mot + ("ERREUR LES POTES")
                Class.DIVIDENDE = Class.DIVIDENDE + mot + "ERREUR"
                Class.SPONSOR = Class.SPONSOR + mot + "ERREUR"
                Class.Site = Class.Site + mot + "ERREUR"
                Class.TICKER = Class.TICKER + mot + "ERREUR"
            
            Class.BLOCDIVIDENDE = Class.BLOCDIVIDENDE + mot + test["Equity"] + " (" + test["Dividende"] + "; code Bloomberg : " + test["Ticker"] +  ";  <sponsor> : "+ test["Sponsor"] +  "; " + test["SiteWeb"] + ")" 

            compteur+=1

